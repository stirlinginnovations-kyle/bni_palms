import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Tuple

from PyPDF2 import PdfReader
from openpyxl import load_workbook

SPREADSHEET_NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
SS_INDEX_ATTR = "{urn:schemas-microsoft-com:office:spreadsheet}Index"
SPREADSHEET_TABLE_START_ROW = 9
REFERRAL_COLUMNS = ("RGI", "RGO")
REFERRALS_TOTAL_COLUMN = "Referrals Total"


def _row_values(row: ET.Element) -> List[str]:
    values: List[str] = []
    col_idx = 1
    for cell in row.findall("ss:Cell", SPREADSHEET_NS):
        idx = cell.get(SS_INDEX_ATTR)
        if idx:
            idx_int = int(idx)
            while col_idx < idx_int:
                values.append("")
                col_idx += 1
        data = cell.find("ss:Data", SPREADSHEET_NS)
        values.append(data.text if data is not None else "")
        col_idx += 1
    return values


def _parse_value(value: str):
    if value is None:
        return ""
    text = str(value).strip()
    if text == "":
        return ""
    if text.endswith("%"):
        num = text[:-1].strip()
        try:
            return float(num)
        except ValueError:
            return text
    cleaned = text.replace(",", "")
    if re.fullmatch(r"\d+", cleaned):
        try:
            return int(cleaned)
        except ValueError:
            return text
    if re.fullmatch(r"\d*\.\d+", cleaned):
        try:
            return float(cleaned)
        except ValueError:
            return text
    return text


def _as_number(value: object) -> float:
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _round_total(value: float) -> object:
    if float(value).is_integer():
        return int(value)
    return round(value, 2)


def tally_referral_columns(rows: List[Dict[str, object]]) -> Dict[str, object]:
    totals = {col: 0.0 for col in REFERRAL_COLUMNS}
    for row in rows:
        for col in REFERRAL_COLUMNS:
            totals[col] += _as_number(row.get(col))
    return {col: _round_total(total) for col, total in totals.items()}


def _row_referrals_total(row: Dict[str, object]) -> object:
    total = 0.0
    for col in REFERRAL_COLUMNS:
        total += _as_number(row.get(col))
    return _round_total(total)


def parse_spreadsheetml_xls(path: Path) -> List[Dict[str, object]]:
    tree = ET.parse(path)
    root = tree.getroot()
    worksheet = root.find("ss:Worksheet", SPREADSHEET_NS)
    if worksheet is None:
        return []
    table = worksheet.find("ss:Table", SPREADSHEET_NS)
    if table is None:
        return []

    rows = table.findall("ss:Row", SPREADSHEET_NS)
    header: List[str] = []
    data_rows: List[Dict[str, object]] = []
    row_num = 0

    for row in rows:
        idx_attr = row.get(SS_INDEX_ATTR)
        if idx_attr:
            try:
                row_num = int(idx_attr)
            except ValueError:
                row_num += 1
        else:
            row_num += 1

        if row_num < SPREADSHEET_TABLE_START_ROW:
            continue

        values = _row_values(row)
        if row_num == SPREADSHEET_TABLE_START_ROW:
            if "First Name" in values and "Last Name" in values:
                header = [v.strip() if v else "" for v in values]
            else:
                # PALMS weekly/YTD files always have table headers on row 9.
                return []
            continue

        if not header:
            continue

        if not any(v not in ("", None) for v in values):
            continue

        first_cell = (values[0] or "").strip()
        if first_cell in {"Visitors", "BNI", "Total"}:
            break

        row_dict: Dict[str, object] = {}
        for idx, col in enumerate(header):
            if not col:
                continue
            cell_value = values[idx] if idx < len(values) else ""
            if col in {"First Name", "Last Name"}:
                row_dict[col] = str(cell_value).strip()
            else:
                row_dict[col] = _parse_value(cell_value)
        row_dict[REFERRALS_TOTAL_COLUMN] = _row_referrals_total(row_dict)
        data_rows.append(row_dict)

    return data_rows


def _parse_excel_cell(value: object):
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return _parse_value(str(value))


def parse_openxml_xlsx(path: Path) -> List[Dict[str, object]]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0] if workbook.worksheets else None
        if worksheet is None:
            return []

        header: List[str] = []
        data_rows: List[Dict[str, object]] = []

        row_iter = worksheet.iter_rows(
            min_row=SPREADSHEET_TABLE_START_ROW,
            values_only=True,
        )
        for offset, row_values in enumerate(row_iter):
            row_num = SPREADSHEET_TABLE_START_ROW + offset
            values = list(row_values or [])

            if row_num == SPREADSHEET_TABLE_START_ROW:
                header = [str(v).strip() if v is not None else "" for v in values]
                if "First Name" not in header or "Last Name" not in header:
                    return []
                continue

            if not header:
                continue

            if not any(v not in ("", None) for v in values):
                continue

            first_cell = str(values[0]).strip() if values and values[0] is not None else ""
            if first_cell in {"Visitors", "BNI", "Total"}:
                break

            row_dict: Dict[str, object] = {}
            for idx, col in enumerate(header):
                if not col:
                    continue
                cell_value = values[idx] if idx < len(values) else None
                if col in {"First Name", "Last Name"}:
                    row_dict[col] = str(cell_value or "").strip()
                else:
                    row_dict[col] = _parse_excel_cell(cell_value)
            row_dict[REFERRALS_TOTAL_COLUMN] = _row_referrals_total(row_dict)
            data_rows.append(row_dict)

        return data_rows
    finally:
        workbook.close()


def parse_chapter_spreadsheet(path: Path) -> List[Dict[str, object]]:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return parse_openxml_xlsx(path)

    if ext == ".xls":
        # Primary format in current workflow.
        rows = parse_spreadsheetml_xls(path)
        if rows:
            return rows
        # Some users rename/export modern Excel files with .xls extension.
        try:
            return parse_openxml_xlsx(path)
        except Exception:
            return []

    if ext == ".xlsm":
        return parse_openxml_xlsx(path)

    return []


TL_COLUMNS = [
    "S",
    "ML",
    "A",
    "P",
    "Wks",
    "TYFCB",
    "CEUs",
    "Points",
    "Given",
    "Recd",
    "121",
    "Vis",
    "Referrals",
    "AttendancePct",
    "Attendance",
    "ReferralsPts",
    "ReferralsAPW",
    "CEUsPts",
    "CEUsAPW",
    "121Pts",
    "121APW",
    "VisitorsPts",
    "VisitorsAPW",
]
TL_POINTS_INDEX_LEGACY = TL_COLUMNS.index("Points")
TL_POINTS_INDEX_CURRENT = 4
TL_EXPECTED_NUMERIC_FIELDS = 23


def _traffic_numeric_value(token: str) -> float | None:
    value = _parse_value(token)
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _percentile(values: List[float], pct: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    idx = int((len(ordered) - 1) * pct)
    return ordered[idx]


def _select_traffic_points_index(numeric_rows: List[List[str]]) -> int:
    # We support two known extraction layouts:
    # - current PDF text layout: points is the 5th numeric token after member name
    # - legacy layout: points aligns to the historical "Points" column index
    candidates = [TL_POINTS_INDEX_CURRENT, TL_POINTS_INDEX_LEGACY]
    best_idx = TL_POINTS_INDEX_CURRENT
    best_score: Tuple[float, float] = (-1.0, -1.0)

    for idx in candidates:
        values: List[float] = []
        for nums in numeric_rows:
            if idx >= len(nums):
                continue
            number = _traffic_numeric_value(nums[idx])
            if number is not None:
                values.append(number)
        if not values:
            continue

        in_range = [v for v in values if 0.0 <= v <= 120.0]
        in_range_ratio = len(in_range) / len(values)
        p90 = _percentile(in_range or values, 0.9)
        score = (in_range_ratio, p90)
        if score > best_score:
            best_score = score
            best_idx = idx

    return best_idx


def parse_traffic_lights_pdf(path: Path) -> List[Dict[str, object]]:
    reader = PdfReader(str(path))
    member_rows: List[Tuple[str, str, str, List[str]]] = []
    numeric_rows: List[List[str]] = []

    num_re = re.compile(r"^\d[\d,]*$")
    dec_re = re.compile(r"^\d*\.\d+$")
    pct_re = re.compile(r"^\d+%$")

    def is_num(tok: str) -> bool:
        return bool(num_re.match(tok) or dec_re.match(tok) or pct_re.match(tok))

    for page in reader.pages:
        text = page.extract_text() or ""
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        if not lines:
            continue

        header = lines[0]
        m = re.search(r"[A-Z]{2} .+? Region for", header)
        if m:
            chapter = header[: m.start()].strip()
        else:
            chapter = header.split("Region for")[0].strip()

        try:
            start = next(i for i, l in enumerate(lines) if l.startswith("Launched"))
        except StopIteration:
            start = 0

        for line in lines[start + 1 :]:
            if (
                "Chapter Totals" in line
                or line.startswith("Designed and produced")
                or line.startswith("To protect")
                or "Personal Data" in line
            ):
                break
            if "," not in line:
                continue

            tokens = line.split()
            num_tokens: List[str] = []
            i = len(tokens) - 1
            while i >= 0:
                tok = tokens[i]
                if is_num(tok):
                    num_tokens.append(tok)
                    i -= 1
                else:
                    break

            if len(num_tokens) < TL_EXPECTED_NUMERIC_FIELDS:
                continue

            # Keep the rightmost report metrics and ignore any extra numeric token(s)
            # that might appear in a member name (for example business ids).
            num_tokens = num_tokens[:TL_EXPECTED_NUMERIC_FIELDS]
            nums = list(reversed(num_tokens))

            name_tokens = tokens[: len(tokens) - TL_EXPECTED_NUMERIC_FIELDS]
            name = " ".join(name_tokens)
            if "," in name:
                last, first = name.split(",", 1)
                last = last.strip()
                first = first.strip()
            else:
                last = name.strip()
                first = ""

            member_rows.append((chapter, first, last, nums))
            numeric_rows.append(nums)

    points_index = _select_traffic_points_index(numeric_rows)
    rows: List[Dict[str, object]] = []
    for chapter, first, last, nums in member_rows:
        if points_index >= len(nums):
            continue
        score_value = _parse_value(nums[points_index])
        row: Dict[str, object] = {
            "Chapter": chapter,
            "First Name": first,
            "Last Name": last,
            "Score": score_value,
            # Keep legacy key so older normalization code remains compatible.
            "Points": score_value,
        }
        rows.append(row)

    return rows
